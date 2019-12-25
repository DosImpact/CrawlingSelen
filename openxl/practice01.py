"""
엑셀 파일 행 채우기 
"""
import openpyxl as xl

wb = xl.Workbook()
sheet = wb.active
sheet.cell(row=1, column=1).value = '이름'
sheet.cell(row=1, column=2).value = '성'
sheet.cell(row=1, column=3).value = '주소'
wb.save('test3.xlsx')



