import os
import time
import csv
import random
import openpyxl as xl


# Config -------------------------------------------
TIEMINTERVAL = 2
INPUT_FILE_NAME = "pro2_PDF.csv"
INPUT_START_LINE = 127  # pk보다 하나 적게 이어서 시작!!
OUTPUT_FILE_NAME = "pro3_PDF.xlsx"
# ------------------------------------------------------
output_format = []
# csv FileRead--------------------------------------------
wb = xl.Workbook()
sheet = wb.active
sheet.cell(row=1, column=1).value = '이름'
sheet.cell(row=1, column=2).value = '성'
sheet.cell(row=1, column=3).value = '주소'

# -------------------------------------------
with open(INPUT_FILE_NAME, encoding="utf-16", newline='') as csvfile:
    spamreader = csv.reader(csvfile, delimiter=',', quotechar='|')
    for row in spamreader:
        xlrow = []
        if(len(row) == 0):
            continue
        elif(len(row) == 2):
            xlrow.extend([row[0], row[1]])
            pass
        elif(len(row) >= 3):
            xlrow.extend([row[0], row[1], row[2]])
        sheet.append(xlrow)

# -------------------------------------------

# CSV 2 xlxs 이름 성 주소 -------------------------------------------


wb.save(OUTPUT_FILE_NAME)

# -------------------------------------------
